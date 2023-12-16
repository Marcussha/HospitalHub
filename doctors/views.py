from django.shortcuts import render, redirect
from doctors.models import Doctors 
from departments.models import Departments
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from django.db import IntegrityError
from django.contrib import messages
from openpyxl import Workbook
from datetime import datetime 
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
import os
from django.contrib.auth.decorators import user_passes_test
from authrole.custom_context import user_is_admin, user_is_doctor

# Create your views here.
def home(request):
    doctor = Doctors.objects.all()
    return render(request, "admin/doctors/show.html", {'doctors': doctor})


@user_passes_test(lambda u: user_is_admin(u), login_url='login')
def index(request):
    doctor = Doctors.objects.all()
    return render(request, "admin/doctors/index.html", {'doctors': doctor})


@user_passes_test(lambda u: user_is_admin(u), login_url='login')
def create(request):
    if request.method == "POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        position = request.POST.get('position')
        note = request.POST.get('note')
        images = request.FILES.get('images')
        
        department = request.POST.get('department')
        department_instance = Departments.objects.get(departmentid=department)
    
        try:
            fs = FileSystemStorage()
            filename = fs.save(images.name, images)
            
            new_doctor = Doctors.objects.create(
                name=name,
                email=email,
                position=position,
                note=note,
                images=filename,
                department=department_instance,
            )
            return redirect ('/doctors/index')

        except IntegrityError as e:
            if 'unique constraint' in str(e).lower() and 'email' in str(e).lower():
                messages.error(request, 'Email already exists. Please choose a different email.')
            else:
                messages.error(request, 'An error occurred during the creation of the doctor.')

    departments = Departments.objects.all()
    return render(request, 'admin/doctors/create.html', {'departments': departments})


@user_passes_test(lambda u: user_is_admin(u), login_url='login')    
def edit(request,id):
    doctors = Doctors.objects.get(id=id)
    return render(request, 'admin/doctors/edit.html', {'doctors': doctors})


@user_passes_test(lambda u: user_is_admin(u), login_url='login')
def update(request, id):
    if request.method == "POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        position = request.POST.get('position')

        doctor = Doctors.objects.get(id=id)

        new_images = request.FILES.get('images')

        if new_images:
            if doctor.images:
                old_images_path = doctor.images.path
                os.remove(old_images_path)
            doctor.images = new_images


        doctor.name = name
        doctor.email = email
        doctor.position = position
        doctor.save()

        return redirect('/doctors/index')

    doctors = Doctors.objects.get(id=id)
    return render(request, 'admin/doctors/edit.html', {'doctors': doctors})


@user_passes_test(lambda u: user_is_admin(u), login_url='login')    
def clear(request, id):
    doctor = Doctors.objects.get(id=id)
    
    if doctor.appointment_count > 0 or doctor.prescriptions_count > 0:
        error_message = "Cannot delete a doctor."
        return render (request, "error_page.html", {'error_message': error_message})
    
    doctor.delete()
    return redirect('/doctors/index')  

def excel(request):
    doctor = Doctors.objects.all()
    
    wb = Workbook()
    ws = wb.active  
    
    headers = ["Name", "Email", "Position", "Department","Note"]
    
    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title
        
    for row_num, doctor in enumerate(doctor, 4):
        ws.cell(row=row_num, column=1, value=doctor.name)
        ws.cell(row=row_num, column=2, value=doctor.email)
        ws.cell(row=row_num, column=3, value=doctor.position)
        ws.cell(row=row_num, column=4, value=doctor.department.named)
        ws.cell(row=row_num, column=5, value=doctor.note) 
        
        
    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
    response['Content-Disposition'] = f'attachment; filename=doctors_{current_datetime}.xlsx'
    
    return response