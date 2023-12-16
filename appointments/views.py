from django.shortcuts import render, redirect
from django.core.mail import send_mail
from django.contrib.auth.decorators import user_passes_test
from authrole.custom_context import user_is_admin, user_is_doctor
from django.http import HttpResponse
from appointments.models import Appointment
from ministration.models import Ministration
from doctors.models import Doctors
from .filters import AppointmentFilter
from datetime import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from django.db import IntegrityError
from django.contrib import messages
import csv


# Create your views here.
@user_passes_test(lambda u: user_is_admin(u) or user_is_doctor(u), login_url='login')
def index(request):
    appointment = Appointment.objects.all()
    appointment_filter = AppointmentFilter(request.GET, queryset=appointment)
    appointment = appointment_filter.qs

    return render(request, "appointments/index.html", {'appointment': appointment, 'filter': appointment_filter})


def create(request):
    if request.method == "POST":
        fullname = request.POST.get('fullname')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        datebooking = request.POST.get('datebooking')
        note = request.POST.get('note')
        timebooking_hour = request.POST.get('timebooking_hour')
        timebooking_minute = request.POST.get('timebooking_minute')

        # Get the selected Ministration and Doctor instances using their primary keys
        serviceid_id = request.POST.get('serviceid')
        serviceid = Ministration.objects.get(minisid=serviceid_id)
        docid_id = request.POST.get('docid')
        docid = Doctors.objects.get(id=docid_id)

        # Combine hour and minute into a valid time object
        timebooking = time(int(timebooking_hour), int(timebooking_minute))

        try:
            Appointment.objects.create(
                fullname=fullname,
                email=email,
                phone=phone,
                datebooking=datebooking,
                serviceid=serviceid,
                docid=docid,
                note=note,
                timebooking=timebooking
            )

            # Send confirmation email
            subject = "Appointment Confirmation"
            message = f"Hello {fullname},\n\nYour appointment is confirmed for {datebooking} at {timebooking_hour}:{timebooking_minute}.\nService: {serviceid}\nDoctor: {docid}\n\nThank you for choosing us!"

            from_email = "thonghqgcs200763@fpt.edu.vn"
            recipient_list = [email]
            send_mail(subject, message, from_email, recipient_list, fail_silently=False)

            messages.success(request, 'Booking successfully')
            return redirect('/appointments/create')

        except IntegrityError as e:
            if 'phone' in str(e):
                messages.error(request, 'Error: Out of range value for phone number')
            else:
                messages.error(request, f'Error: {e}')

    services = Ministration.objects.all()
    doctors = Doctors.objects.all()

    return render(request, "appointments/create.html", {'services': services, 'doctors': doctors})


@user_passes_test(lambda u: user_is_admin(u) or user_is_doctor(u), login_url='login')
def clear (request,id):
    appointment = Appointment.objects.get( appid =id)
    appointment.delete()
    return redirect('/appointments') 


@user_passes_test(lambda u: user_is_admin(u) or user_is_doctor(u), login_url='login')
def delete (request,id):
    appointment = Appointment.objects.get( appid =id)
    appointment.delete()
    return redirect('/appointments/home')


@user_passes_test(lambda u: user_is_admin(u) or user_is_doctor(u), login_url='login')
def export_filtered_excel(request):
    appointments = AppointmentFilter(request.GET, queryset=Appointment.objects.all()).qs
    
    wb = Workbook()
    ws = wb.active

    headers = ["Full Name", "Email", "Phone", "Date", "Time", "Service", "Doctor", "Notes"]

    # Set the column headers in the worksheet
    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title

    # Populate the data rows
    for row_num, appointment in enumerate(appointments, 2):  
        ws.cell(row=row_num, column=1, value=appointment.fullname)
        ws.cell(row=row_num, column=2, value=appointment.email)
        ws.cell(row=row_num, column=3, value=appointment.phone)
        ws.cell(row=row_num, column=4, value=appointment.datebooking)
        ws.cell(row=row_num, column=5, value=appointment.timebooking)
        ws.cell(row=row_num, column=6, value=appointment.serviceid.name_ministration)  
        ws.cell(row=row_num, column=7, value=appointment.docid.name)   
        ws.cell(row=row_num, column=8, value=appointment.note)

    # Create a response with the Excel file
    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Set the filename for the download
    current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
    response['Content-Disposition'] = f'attachment; filename=appointments_{current_datetime}.xlsx'

    return response


@user_passes_test(lambda u: user_is_admin(u) or user_is_doctor(u), login_url='login')
def export_csv(request):
    appointments = Appointment.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="appointments.csv"'

    writer = csv.writer(response)
    writer.writerow(["Full Name", "Email", "Phone", "Date", "Time", "Service", "Doctor", "Notes"])

    for appointment in appointments:
        writer.writerow([
            appointment.fullname,
            appointment.email,
            appointment.phone,
            appointment.datebooking,
            appointment.timebooking,
            appointment.serviceid,
            appointment.docid,
            appointment.note,
        ])

    return response


### View for admin page
@user_passes_test(lambda u: user_is_admin(u), login_url='login')
def home(request):
    appointment = Appointment.objects.all()
    appointment_filter = AppointmentFilter(request.GET, queryset=appointment)
    appointment = appointment_filter.qs
    return render(request, "admin/appointment/home.html", {'appointment': appointment, 'filter': appointment_filter})


@user_passes_test(lambda u: user_is_admin(u) or user_is_doctor(u), login_url='login')
def export_excel(request):
    appointments = Appointment.objects.all()

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define the headers
    headers = ["Full Name", "Email", "Phone", "Date", "Time", "Service", "Doctor", "Notes"]

    # Set the column headers in the worksheet
    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title

    # Populate the data rows
    for row_num, appointment in enumerate(appointments, 2):  
        ws.cell(row=row_num, column=1, value=appointment.fullname)
        ws.cell(row=row_num, column=2, value=appointment.email)
        ws.cell(row=row_num, column=3, value=appointment.phone)
        ws.cell(row=row_num, column=4, value=appointment.datebooking)
        ws.cell(row=row_num, column=5, value=appointment.timebooking)
        ws.cell(row=row_num, column=6, value=appointment.serviceid.name_ministration)  
        ws.cell(row=row_num, column=7, value=appointment.docid.name)   
        ws.cell(row=row_num, column=8, value=appointment.note)

    # Create a response with the Excel file
    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Set the filename for the download
    current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
    response['Content-Disposition'] = f'attachment; filename=appointments_{current_datetime}.xlsx'

    return response
